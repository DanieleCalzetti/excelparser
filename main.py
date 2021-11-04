from tkinter.constants import END
from openpyxl import load_workbook
import sys
import tkinter as tk
from tkinter.constants import END
from tkinter.filedialog import askopenfilename

sys.path.append(".")

from read_document import read_document 

# TODO:
# Understand classes 
# Understand methods vs functions
# Understand importing from other files 

ws = tk.Tk()
ws.title('Excel Parser')
ws.geometry('900x900') 

top_title = tk.Frame(ws)
values = tk.Frame(ws)
middle_title = tk.Frame(ws)
results = tk.Frame(ws)

ws.grid_rowconfigure(3, weight=1)
ws.grid_columnconfigure(0, weight=1)

top_title.grid(row=0, sticky="")
values.grid(row=1, sticky="N")
middle_title.grid(row=2, sticky="N")
results.grid(row=3, sticky="N")

# Top title GUI

title=tk.Label(top_title,text="Insert the column value", width=20, pady=10, font=("bold",20))
title.grid(row=0, column=0)

# Values GUI

coverage_label=tk.Label(values,text="Coverage", anchor="w", width=22, pady=5, font=("bold",10))
coverage_label.grid(row=0, column=0)

coverage_entry=tk.Entry(values)
coverage_entry.insert(END, 'Copertura')
coverage_entry.grid(row=0, column=1)

impression_label=tk.Label(values,text="Impression", anchor="w", width=22, pady=5, font=("bold",10))
impression_label.grid(row=1, column=0)

impression_entry=tk.Entry(values)
impression_entry.insert(END, 'Impression')
impression_entry.grid(row=1, column=1 )

likes_label=tk.Label(values,text="Likes", anchor="w", width=22, pady=5, font=("bold",10))
likes_label.grid(row=2, column=0)

likes_entry=tk.Entry(values)
likes_entry.insert(END, 'Likes')
likes_entry.grid(row=2, column=1 )

hashtag_label=tk.Label(values,text="Hashtag", anchor="w", width=22, pady=5, font=("bold",10))
hashtag_label.grid(row=3, column=0)

hashtag_entry=tk.Entry(values)
hashtag_entry.insert(END, 'Hashtag')
hashtag_entry.grid(row=3, column=1 )

first_post_label=tk.Label(values,text="Row of the first post", anchor="w", width=22, pady=5, font=("bold",10))
first_post_label.grid(row=4, column=0)

first_post_entry=tk.Entry(values)
first_post_entry.insert(END, 10)
first_post_entry.grid(row=4, column=1 )

header_row_label=tk.Label(values,text="Header row", anchor="w", width=22, pady=5, font=("bold",10))
header_row_label.grid(row=5, column=0)

header_row_entry=tk.Entry(values)
header_row_entry.insert(END, 8)
header_row_entry.grid(row=5, column=1)

excel_label = tk.Label(values, text='Upload Excel file', anchor="w", width=22, pady=5, font=("bold",10))
excel_label.grid(row=6, column=0)

excel_button = tk.Button(values, text ='Choose File and get column values', command = lambda:open_file())
excel_button.grid(row=6, column=1)

# Middle title GUI

results_title = tk.Label(middle_title,text="Values returned", width=20, pady=10, font=("bold",20))
results_title.grid(row=0, column=0)

# Results GUI

coordinate_top_three_cov = tk.Label(results, text = 'The three posts with highest coverage are:', anchor="w", width=66, pady=5, font=("bold",10))
coordinate_top_three_cov.grid(row=0, column=0)

coordinate_top_three_cov_entry=tk.Entry(results)
coordinate_top_three_cov_entry.insert(END, "")
coordinate_top_three_cov_entry.grid(row=0, column=1)

coordinate_top_three_like = tk.Label(results, text = 'The three posts with highest likes are:', anchor="w", width=66, pady=5, font=("bold",10))
coordinate_top_three_like.grid(row=1, column=0)

coordinate_top_three_like=tk.Entry(results)
coordinate_top_three_like.insert(END, "")
coordinate_top_three_like.grid(row=1, column=1)

coordinate_top_three_impre = tk.Label(results, text = 'The three posts with highest impression are:', anchor="w", width=66, pady=5, font=("bold",10))
coordinate_top_three_impre.grid(row=2, column=0)

coordinate_top_three_impre=tk.Entry(results)
coordinate_top_three_impre.insert(END, "")
coordinate_top_three_impre.grid(row=2, column=1)

common_hashtag_cov = tk.Label(results, text = 'The commons hashtags for the top three with highest coverage are:', anchor="w", width=66, pady=5, font=("bold",10))
common_hashtag_cov.grid(row=3, column=0)

common_hashtag_cov=tk.Entry(results)
common_hashtag_cov.insert(END, "")
common_hashtag_cov.grid(row=3, column=1)

common_hashtag_like = tk.Label(results, text = 'The commons hashtags for the top three with highest likes are:', anchor="w", width=66, pady=5, font=("bold",10))
common_hashtag_like.grid(row=4, column=0)

common_hashtag_like=tk.Entry(results)
common_hashtag_like.insert(END, "")
common_hashtag_like.grid(row=4, column=1)

common_hashtag_impre = tk.Label(results, text = 'The commons hashtags for the top three with highest impressions are:', anchor="w", width=66, pady=5, font=("bold",10))
common_hashtag_impre.grid(row=5, column=0)

common_hashtag_impre=tk.Entry(results)
common_hashtag_impre.insert(END, "")
common_hashtag_impre.grid(row=5, column=1)

common_hashtag_label = tk.Label(results, text = 'The commons hashtags are:', anchor="w", width=66, pady=5, font=("bold",10))
common_hashtag_label.grid(row=6, column=0)

common_hashtag_label=tk.Entry(results)
common_hashtag_label.insert(END, "")
common_hashtag_label.grid(row=6, column=1)

def get_args():
    args = {}
    args["hashtag_header"] = hashtag_entry.get()
    args["coverage_header"] = coverage_entry.get()
    args["like_header"] = likes_entry.get()
    args["impression_header"] = impression_entry.get()
    args["row_header"] = header_row_entry.get()
    args["first_post"] = first_post_entry.get() 
    return args

def open_file():
    file_path = askopenfilename(filetypes=[("Excel files", "*.xlsx")]) 
    coordinate_top_three_cov_entry.delete(0, 'end')
    coordinate_top_three_like.delete(0, 'end')
    coordinate_top_three_impre.delete(0, 'end')
    common_hashtag_cov.delete(0, 'end')
    common_hashtag_like.delete(0, 'end')
    common_hashtag_impre.delete(0, 'end')
    common_hashtag_label.delete(0, 'end') 
    if file_path != "":
        wb = load_workbook(file_path)
        sheets = wb.sheetnames
        sheet = wb[sheets[0]]
        args = get_args()
        s = read_document(sheet)
        headers = s.find_header(args["coverage_header"], args["like_header"], args["row_header"], args["impression_header"], args["hashtag_header"])
        if "" in headers:
            coordinate_top_three_cov_entry.insert(0, "Error, please review your headers!")
            coordinate_top_three_like.insert(0, "Error, please review your headers!")
            coordinate_top_three_impre.insert(0, "Error, please review your headers!")
            common_hashtag_cov.insert(0, "Error, please review your headers!")
            common_hashtag_like.insert(0, "Error, please review your headers!")
            common_hashtag_impre.insert(0, "Error, please review your headers!")
            common_hashtag_label.insert(0, "Error, please review your headers!")            
        else:
            s.check_row_used(args["first_post"])
            top_posts = s.find_top_post()
            hashtag_top_three = s.compare_hashtag_top_three()
            common_global_hashtag = s.compare_hashtag()
            top_posts_cov_list = list(top_posts[0])
            top_posts_like_list = list(top_posts[1])
            top_posts_impre_list = list(top_posts[2])

            coordinate_top_three_cov_entry.insert(0, "".join(["A", top_posts_cov_list[0], ", ", "A", top_posts_cov_list[1], ", ", "A", top_posts_cov_list[2]]))
            coordinate_top_three_like.insert(0,  "".join(["A", top_posts_like_list[0], ", ", "A", top_posts_like_list[1], ", ", "A", top_posts_like_list[2]]))
            coordinate_top_three_impre.insert(0, "".join(["A", top_posts_impre_list[0], ", ", "A", top_posts_impre_list[1], ", ", "A", top_posts_impre_list[2]]))

            common_hashtag_cov.insert(0, " ".join(hashtag_top_three[0]))
            common_hashtag_like.insert(0, " ".join(hashtag_top_three[1]))
            common_hashtag_impre.insert(0, " ".join(hashtag_top_three[2]))
            common_hashtag_label.insert(0, common_global_hashtag)
    else:
        pass


ws.mainloop()